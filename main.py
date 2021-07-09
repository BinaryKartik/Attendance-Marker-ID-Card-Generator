from PIL import Image, ImageFont, ImageDraw
import cv2
import qrcode
import openpyxl
import datetime
import os
from openpyxl import Workbook
from pathlib import Path
a = int(input("Type 1 to Mark Attendance, 2 to Generate ID Card and 3 to Clear Data: "))
y = 0
if a == 1:
    if not os.path.exists("Data for ID Generator.xlsx"):
        xlfile = Workbook()
        xlfile.save("Data for ID Generator.xlsx")
        xlfile.create_sheet("Data")
        xlfile.create_sheet("Background Tasks")
        sh1 = xlfile["Data"]
        sh2 = xlfile["Background Tasks"]
        sh1.cell(row=1, column=1, value="Name")
        sh1.cell(row=1, column=2, value="DOB")
        sh1.cell(row=1, column=3, value="Class")
        sh1.cell(row=1, column=4, value="School")
        sh2.cell(row=1, column=1, value=1)
        sh2.cell(row=1, column=2, value=3)
        xlfile.remove(xlfile["Sheet"])
        xlfile.save("Data for ID Generator.xlsx")
    newfile = openpyxl.load_workbook("Data for ID Generator.xlsx")
    newfile.create_sheet("Sheet1")
    sh1 = newfile["Sheet1"]
    sh2 = newfile["Background Tasks"]
    img = input("Name of image of the ID card:")
    im = cv2.imread(img)
    detectqr = cv2.QRCodeDetector()
    data1, bbox, straight_qrcode = detectqr.detectAndDecode(im)
    time = datetime.datetime.now().strftime('%d-%m-%Y')
    row = sh1.max_row
    column = sh1.max_column
    for i in range(1, row + 1):
        for j in range(1, column + 1):
            if sh1.cell(i, j).value == data1:
                y = 1
                k = sh2.cell(row=1, column=2).value + 1
                sh2.cell(row=1, column=2, value=k)
                sh1.cell(row=i, column=j+k, value="Present")
                sh1.cell(row=1, column=j+k, value=time)
                newfile.save("Data for ID Generator.xlsx")
                print("Your Attendance has been Marked.")
    if y == 0:
        print("Please Register yourself first by making an ID card.")
elif a == 2:
    if not os.path.exists("Data for ID Generator.xlsx"):
        xlfile = Workbook()
        xlfile.save("Data for ID Generator.xlsx")
        xlfile.create_sheet("Sheet1")
        xlfile.create_sheet("Background Tasks")
        sh1 = xlfile["Sheet1"]
        sh2 = xlfile["Background Tasks"]
        sh1.cell(row=1, column=1, value="Name")
        sh1.cell(row=1, column=2, value="DOB")
        sh1.cell(row=1, column=3, value="Class")
        sh1.cell(row=1, column=4, value="School")
        sh2.cell(row=1, column=1, value=1)
        sh2.cell(row=1, column=2, value=3)
        xlfile.remove(xlfile["Sheet"])
        xlfile.save("Data for ID Generator.xlsx")
    profilepic = input("Insert Profile Pic by giving name:")
    name = input("Name Of Student: ")
    dob = input("Date Of Birth: ")
    sclass = input("Class and Section: ")
    school = input("Name of School: ")
    data2 = name
    img1 = Image.open("ID Prototype.jpg")
    font = ImageFont.truetype("JosefinSans-Regular.ttf", 26)
    draw = ImageDraw.Draw(img1)
    draw.text((387, 165), name, (0,0,0), font=font)
    draw.text((387, 221), dob, (0, 0, 0), font=font)
    draw.text((387, 277), sclass, (0, 0, 0), font=font)
    draw.text((387, 329), school, (0, 0, 0), font=font)
    filename = "qrcode.jpg"
    img = qrcode.make(data2)
    img.save(filename)
    img = Image.open("qrcode.jpg")
    img = img.resize((105, 105))
    img1.paste(img, (648, 123))
    img3 = Image.open(profilepic)
    img3 = img3.resize((189, 208))
    img1.paste(img3, (67, 156))
    img1.save(f"{name}'s ID CARD.png")
    file = openpyxl.load_workbook("Data for ID Generator.xlsx")
    sh1 = file["Sheet1"]
    sh2 = file["Background Tasks"]
    recall = int(sh2.cell(row=1, column=1).value) + 1
    for i in range(1, 5):
        if i == 1:
            sh1.cell(row=recall, column=1, value=name)
        elif i == 2:
            sh1.cell(row=recall, column=2, value=dob)
        elif i == 3:
            sh1.cell(row=recall, column=3, value=sclass)
        else:
            sh1.cell(row=recall, column=4, value=school)
    sh2.cell(row=1, column=1, value=recall)
    file.save("Data for ID Generator.xlsx")
elif a == 3:
    if not os.path.exists("Data for ID Generator.xlsx"):
        xlfile = Workbook()
        xlfile.save("Data for ID Generator.xlsx")
        xlfile.create_sheet("Sheet1")
        xlfile.create_sheet("Background Tasks")
        sh1 = xlfile["Sheet1"]
        sh2 = xlfile["Background Tasks"]
        sh1.cell(row=1, column=1, value="Name")
        sh1.cell(row=1, column=2, value="DOB")
        sh1.cell(row=1, column=3, value="Class")
        sh1.cell(row=1, column=4, value="School")
        sh2.cell(row=1, column=1, value=1)
        sh2.cell(row=1, column=2, value=3)
        xlfile.remove(xlfile["Sheet"])
        xlfile.save("Data for ID Generator.xlsx")
    file = openpyxl.load_workbook("Data for ID Generator.xlsx")
    sh1 = file["Sheet1"]
    sh2 = file["Background Tasks"]
    row = sh1.max_row
    column = sh1.max_column
    for m in range(1, row + 1):
        for n in range(1, column + 1):
            sh1.cell(row=m, column=n, value="")
    sh1.cell(row=1, column=1, value="Name")
    sh1.cell(row=1, column=2, value="DOB")
    sh1.cell(row=1, column=3, value="Class")
    sh1.cell(row=1, column=4, value="School")
    sh2.cell(row=1, column=1, value=1)
    sh2.cell(row=1, column=2, value=3)
    file.save("Data for ID Generator.xlsx")
    print("Data Cleared")
else:
    print("Please enter 1, 2 or 3")
